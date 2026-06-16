"""
KreditLab Excel Export Module
Sheets: Summary | P&L | Balance Sheet | Ratios | Working Capital | DSCR & Funding | TNW | Observations
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict

C = {
    'hdr_bg': '1E3A5F', 'hdr_ft': 'FFFFFF', 'sec_bg': 'E8F0FE', 'sec_ft': '1D4ED8',
    'tot_bg': 'E6F4EA', 'tot_ft': '047857', 'gtot_bg': 'DBEAFE', 'gtot_ft': '1E3A5F',
    'err_bg': 'FEE2E2', 'err_ft': 'B91C1C', 'wrn_ft': 'B45309',
    'ebi_bg': 'F3E8FF', 'ebi_ft': '6D28D9', 'alt': 'F8FAFC', 'bdr': 'D1D5DB',
}
BDR = Border(left=Side('thin', color=C['bdr']), right=Side('thin', color=C['bdr']),
             top=Side('thin', color=C['bdr']), bottom=Side('thin', color=C['bdr']))
NF = '#,##0;(#,##0);"-"'

def _ft(bold=False, sz=10, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=sz, color=color, italic=italic)
def _fl(c): return PatternFill('solid', fgColor=c)
def _al(h='left', wrap=True): return Alignment(horizontal=h, vertical='center', wrap_text=wrap)
def _widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
def _hdr_row(ws, row, vals):
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = _ft(True, 9, C['hdr_ft']); c.fill = _fl(C['hdr_bg'])
        c.alignment = _al('center' if col > 1 else 'left'); c.border = BDR
def _sec_row(ws, row, text, ncols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = _ft(True, 10, C['sec_ft']); c.fill = _fl(C['sec_bg']); c.border = BDR
    for i in range(2, ncols+1):
        x = ws.cell(row=row, column=i); x.fill = _fl(C['sec_bg']); x.border = BDR
def _row(ws, row, vals, bold=False, bg=None, fc='000000', nfmt=None, indent=0):
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = _ft(bold, 10, fc)
        if bg: c.fill = _fl(bg)
        c.alignment = _al('right' if col > 1 and isinstance(v, (int, float)) else 'left')
        c.border = BDR
        if nfmt and col > 1 and isinstance(v, (int, float)): c.number_format = nfmt
        if indent and col == 1:
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=indent)
def _gv(obj, pks):
    if isinstance(obj, dict):
        v = obj.get('values', obj)
        return [v.get(pk, 0) for pk in pks]
    return [0]*len(pks)
def _pks(data): return list(data.get('company_info', {}).get('periods_analyzed', {}).keys())
def _labels(data): return list(data.get('company_info', {}).get('periods_analyzed', {}).values())

def _summary(wb, data, pks, labels):
    ws = wb.active; ws.title = "Summary"; _widths(ws, [30, 50])
    ci = data.get('company_info', {}); si = data.get('_schema_info', {})
    ws.cell(1, 1, "FINANCIAL STATEMENT ANALYSIS").font = _ft(True, 14, C['hdr_bg'])
    ws.cell(2, 1, ci.get('legal_name', ci.get('name', ''))).font = _ft(True, 12, C['sec_ft'])
    r = 4
    for lbl, val in [("Registration No.", ci.get('registration_no','')), ("Principal Activities", ci.get('principal_activities','')),
                     ("Financial Year End", ci.get('financial_year_end','')), ("Analysis Basis", si.get('analysis_basis','')),
                     ("Currency", si.get('currency_unit','RM')), ("Generated", si.get('generation_date',''))]:
        ws.cell(r, 1, lbl).font = _ft(True); ws.cell(r, 2, val).font = _ft(); r += 1
    r += 1; ws.cell(r, 1, "PERIODS ANALYZED").font = _ft(True, 11, C['sec_ft']); r += 1
    for pk, lb in zip(pks, labels): ws.cell(r, 1, pk).font = _ft(); ws.cell(r, 2, lb).font = _ft(); r += 1
    r += 1; ws.cell(r, 1, "AUDIT OPINIONS").font = _ft(True, 11, C['sec_ft']); r += 1
    for pk in pks:
        ao = ci.get('audit_opinion', {}).get(pk, {})
        if ao:
            ws.cell(r, 1, pk).font = _ft(True)
            ws.cell(r, 2, f"{ao.get('opinion_type','N/A')} - {ao.get('auditor_name','N/A')}").font = _ft(); r += 1
    directors = ci.get('directors', [])
    if directors:
        r += 1; ws.cell(r, 1, "DIRECTORS").font = _ft(True, 11, C['sec_ft']); r += 1
        for d in directors: ws.cell(r, 1, d).font = _ft(); r += 1

def _pl(wb, data, pks, labels):
    ws = wb.create_sheet("P&L"); nc = 1+len(pks); _widths(ws, [45]+[18]*len(pks))
    _hdr_row(ws, 1, ["DESCRIPTION"]+[l.upper() for l in labels])
    inc = data.get('statement_of_comprehensive_income', {}); r = 2
    def _sec(name, sd, r, exp=False):
        if not sd: return r
        _sec_row(ws, r, name.upper(), nc); r += 1
        for k, it in sd.get('line_items', {}).items():
            dn = it.get('display_name', k) if isinstance(it, dict) else k
            vs = _gv(it, pks) if isinstance(it, dict) else [0]*len(pks)
            _row(ws, r, [dn]+vs, indent=1, nfmt=NF); r += 1
        t = sd.get('total', {})
        if t:
            bg = C['err_bg'] if exp else C['tot_bg']; fc = C['err_ft'] if exp else C['tot_ft']
            _row(ws, r, [t.get('display_name','Total')]+_gv(t,pks), True, bg, fc, NF); r += 1
        return r
    r = _sec("Revenue", inc.get('revenue',{}), r)
    r = _sec("Cost of Sales", inc.get('cost_of_sales',{}), r, True)
    gp = inc.get('gross_profit', {})
    _row(ws, r, [gp.get('display_name','Gross Profit')]+_gv(gp,pks), True, C['tot_bg'], C['tot_ft'], NF); r += 1
    r = _sec("Other Income", inc.get('other_income',{}), r)
    for ck, cd in inc.get('operating_expenses', {}).items():
        if isinstance(cd, dict) and 'line_items' in cd:
            r = _sec(cd.get('total',{}).get('display_name',ck), cd, r, True)
    op = inc.get('operating_profit', {})
    _row(ws, r, [op.get('display_name','Operating Profit')]+_gv(op,pks), True, C['gtot_bg'], C['gtot_ft'], NF); r += 1
    r = _sec("Finance Costs", inc.get('finance_costs',{}), r, True)
    pbt = inc.get('profit_before_tax', {})
    _row(ws, r, [pbt.get('display_name','PBT')]+_gv(pbt,pks), True, C['gtot_bg'], C['gtot_ft'], NF); r += 1
    r = _sec("Taxation", inc.get('taxation',{}), r, True)
    npat = inc.get('net_profit_after_tax', {})
    _row(ws, r, [npat.get('display_name','NPAT')]+_gv(npat,pks), True, C['gtot_bg'], C['gtot_ft'], NF); r += 1
    ebitda = inc.get('ebitda', {})
    _row(ws, r, [ebitda.get('display_name','EBITDA')]+_gv(ebitda,pks), True, C['ebi_bg'], C['ebi_ft'], NF)

def _bs(wb, data, pks, labels):
    ws = wb.create_sheet("Balance Sheet"); nc = 1+len(pks); _widths(ws, [45]+[18]*len(pks))
    _hdr_row(ws, 1, ["DESCRIPTION"]+[l.upper() for l in labels])
    bs = data.get('statement_of_financial_position', {}); r = 2
    def _bsec(name, sd, r):
        if not sd: return r
        _sec_row(ws, r, name.upper(), nc); r += 1
        for k, it in sd.items():
            if k == 'total': continue
            if isinstance(it, dict):
                if 'line_items' in it:
                    for sk, si in it.get('line_items', {}).items():
                        sn = si.get('display_name', sk) if isinstance(si, dict) else sk
                        _row(ws, r, [sn]+(_gv(si,pks) if isinstance(si,dict) else [0]*len(pks)), indent=2, nfmt=NF); r += 1
                    tt = it.get('total', {})
                    _row(ws, r, [tt.get('display_name',k)]+_gv(tt,pks), True, C['alt'], nfmt=NF, indent=1); r += 1
                elif 'values' in it:
                    _row(ws, r, [it.get('display_name',k)]+_gv(it,pks), indent=1, nfmt=NF); r += 1
        t = sd.get('total', {})
        if t: _row(ws, r, [t.get('display_name','Total')]+_gv(t,pks), True, C['tot_bg'], C['tot_ft'], NF); r += 1
        return r
    r = _bsec("Non-Current Assets", bs.get('non_current_assets',{}), r)
    r = _bsec("Current Assets", bs.get('current_assets',{}), r)
    ta = bs.get('total_assets', {})
    _row(ws, r, [ta.get('display_name','Total Assets')]+_gv(ta,pks), True, C['gtot_bg'], C['gtot_ft'], NF); r += 2
    r = _bsec("Equity", bs.get('equity',{}), r)
    r = _bsec("Non-Current Liabilities", bs.get('non_current_liabilities',{}), r)
    r = _bsec("Current Liabilities", bs.get('current_liabilities',{}), r)
    tl = bs.get('total_liabilities', {})
    _row(ws, r, [tl.get('display_name','Total Liabilities')]+_gv(tl,pks), True, C['err_bg'], C['err_ft'], NF); r += 1
    tel = bs.get('total_equity_and_liabilities', {})
    _row(ws, r, [tel.get('display_name','Total E+L')]+_gv(tel,pks), True, C['gtot_bg'], C['gtot_ft'], NF); r += 2
    ic = data.get('integrity_check', {}).get('balance_sheet_verification', {})
    _sec_row(ws, r, "BALANCE SHEET VERIFICATION", nc); r += 1
    for pk, lb in zip(pks, labels):
        chk = ic.get(pk, {})
        st = "Balanced" if chk.get('balanced', False) else f"Variance: {chk.get('variance','?')}"
        _row(ws, r, [lb, st]+[None]*(len(pks)-1)); r += 1

def _ratios(wb, data, pks, labels):
    ws = wb.create_sheet("Ratios"); nc = 2+len(pks); _widths(ws, [28]+[16]*len(pks)+[12])
    _hdr_row(ws, 1, ["RATIO"]+[l.upper() for l in labels]+["BENCHMARK"])
    ratios = data.get('financial_ratios', {}); r = 2
    for ck, cd in ratios.items():
        _sec_row(ws, r, ck.replace('_',' ').title(), nc); r += 1
        for rk, rd in cd.items():
            if not isinstance(rd, dict): continue
            nm = rd.get('display_name', rk); u = rd.get('unit',''); bm = rd.get('benchmark','')
            vs = _gv(rd, pks)
            fmt = '0.00"%"' if u=='%' else '0.00"x"' if u=='x' else '0 "days"' if u=='days' else '0.00'
            _row(ws, r, [nm]+vs+[bm], nfmt=fmt); r += 1
            pa = rd.get('values_period_adjusted', {})
            if pa and any(pa.get(pk) != rd.get('values_standard',{}).get(pk) for pk in pks if pa.get(pk)):
                pa_vals = [pa.get(pk, '') for pk in pks]
                _row(ws, r, ["  Period-Adjusted"]+pa_vals+[''], indent=1, fc=C['wrn_ft'], nfmt=fmt); r += 1

def _wc(wb, data, pks, labels):
    ws = wb.create_sheet("Working Capital"); nc = 1+len(pks); _widths(ws, [45]+[18]*len(pks))
    _hdr_row(ws, 1, ["METRIC"]+[l.upper() for l in labels])
    wca = data.get('working_capital_analysis', {}); r = 2
    owc = wca.get('operating_working_capital', {})
    _sec_row(ws, r, "OPERATING WORKING CAPITAL", nc); r += 1
    _row(ws, r, ["OWC (Trade Rec + Inventory - Trade Pay)"]+_gv(owc,pks), True, nfmt=NF); r += 1
    for ck, cv in owc.get('components', {}).items():
        cn = ck.replace('_',' ').title()
        cvs = [cv.get(pk,0) for pk in pks] if isinstance(cv, dict) else [0]*len(pks)
        _row(ws, r, [cn]+cvs, indent=1, nfmt=NF); r += 1
    r += 1; wcr = wca.get('working_capital_requirement', {})
    _sec_row(ws, r, "WORKING CAPITAL REQUIREMENT", nc); r += 1
    _row(ws, r, ["WCR Amount"]+_gv(wcr,pks), True, nfmt=NF); r += 2
    assess = wca.get('working_capital_assessment', {})
    _sec_row(ws, r, "ASSESSMENT", nc); r += 1
    for lbl, val in [("Needs WC Facility", "Yes" if assess.get('needs_wc_facility') else "No"),
                     ("CCC Status", assess.get('ccc_status','')), ("OWC Status", assess.get('owc_status','')),
                     ("Recommended Facility", assess.get('recommended_facility_type','')),
                     ("Recommended Amount", assess.get('recommended_facility_amount', 0))]:
        _row(ws, r, [lbl, val]+[None]*(len(pks)-1)); r += 1
    rat = assess.get('rationale', '')
    if rat:
        r += 1; ws.cell(r, 1, "Rationale:").font = _ft(True); r += 1
        c = ws.cell(r, 1, rat); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        ws.row_dimensions[r].height = 80

def _dscr(wb, data, pks, labels):
    ws = wb.create_sheet("DSCR & Funding"); nc = 1+len(pks); _widths(ws, [45]+[18]*len(pks))
    _hdr_row(ws, 1, ["COMPONENT"]+[l.upper() for l in labels])
    dscr = data.get('dscr_analysis', {}); r = 2
    fc = dscr.get('facility_classification', {})
    _sec_row(ws, r, "FACILITY CLASSIFICATION", nc); r += 1
    tf = fc.get('term_facilities', {}); rv = fc.get('revolving_facilities', {})
    _row(ws, r, ["Term: "+tf.get('description','')]+[None]*len(pks)); r += 1
    for f in tf.get('facilities', []): _row(ws, r, ["  "+f]+[None]*len(pks), indent=1); r += 1
    _row(ws, r, ["Revolving: "+rv.get('description','')]+[None]*len(pks)); r += 1
    for f in rv.get('facilities', []): _row(ws, r, ["  "+f]+[None]*len(pks), indent=1); r += 1
    r += 1; calc = dscr.get('calculation', {})
    _sec_row(ws, r, "DSCR CALCULATION", nc); r += 1
    _row(ws, r, ["EBITDA"]+[calc.get(pk,{}).get('ebitda',0) for pk in pks], nfmt=NF); r += 1
    _row(ws, r, ["Principal Repayment"]+[calc.get(pk,{}).get('debt_service',{}).get('principal_repayment',{}).get('total_principal',0) for pk in pks], nfmt=NF); r += 1
    _row(ws, r, ["Interest Expense"]+[calc.get(pk,{}).get('debt_service',{}).get('interest_expense',0) for pk in pks], nfmt=NF); r += 1
    _row(ws, r, ["Total Debt Service"]+[calc.get(pk,{}).get('debt_service',{}).get('total_debt_service',0) for pk in pks], True, C['tot_bg'], C['tot_ft'], NF); r += 1
    _row(ws, r, ["DSCR"]+[calc.get(pk,{}).get('dscr',0) for pk in pks], True, C['gtot_bg'], C['gtot_ft'], '0.00"x"'); r += 2
    asmt = dscr.get('assessment', '')
    if asmt:
        _sec_row(ws, r, "DSCR ASSESSMENT", nc); r += 1
        c = ws.cell(r, 1, asmt); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        ws.row_dimensions[r].height = 80; r += 2
    fm = data.get('funding_mismatch_analysis', {}); gap = fm.get('layer_1_gap_identification', {})
    _sec_row(ws, r, "FUNDING GAP ANALYSIS", nc); r += 1
    _row(ws, r, ["Non-Current Assets"]+[gap.get(pk,{}).get('non_current_assets',0) for pk in pks], nfmt=NF); r += 1
    _row(ws, r, ["Long-Term Funding"]+[gap.get(pk,{}).get('long_term_funding',{}).get('total',0) for pk in pks], nfmt=NF); r += 1
    _row(ws, r, ["Funding Gap"]+[gap.get(pk,{}).get('funding_gap',0) for pk in pks], True, nfmt=NF); r += 1
    _row(ws, r, ["Gap % of NCA"]+[gap.get(pk,{}).get('gap_as_percentage_of_nca',0) for pk in pks], nfmt='0.00"%"'); r += 1
    _row(ws, r, ["Status"]+[gap.get(pk,{}).get('status','') for pk in pks]); r += 2
    fsa = fm.get('funding_structure_assessment', {})
    _row(ws, r, ["Sustainability Rating", fsa.get('overall_sustainability_rating','')]+[None]*(len(pks)-1), True); r += 1
    for rf in fsa.get('risk_flags', []):
        _row(ws, r, ["Risk: "+rf]+[None]*len(pks), fc=C['wrn_ft']); r += 1
    r += 1; fp = data.get('funding_profile', {}).get('existing_facilities_identified', {})
    _sec_row(ws, r, "EXISTING FACILITIES", nc); r += 1
    fac_hdrs = ["Facility", "Current", "Non-Current", "Total"]
    _hdr_row(ws, r, fac_hdrs[:nc]); r += 1
    for fk, fd in fp.items():
        if fk == 'total_borrowings' or not isinstance(fd, dict): continue
        nm = fk.replace('_',' ').title()
        if 'current_portion' in fd:
            _row(ws, r, [nm, fd.get('current_portion',0), fd.get('non_current_portion',0), fd.get('total',0)][:nc], nfmt=NF)
        elif 'amount' in fd:
            _row(ws, r, [nm, fd.get('amount',0), 0, fd.get('amount',0)][:nc], nfmt=NF)
        r += 1
    _row(ws, r, ["Total Borrowings", None, None, fp.get('total_borrowings',0)][:nc], True, C['tot_bg'], C['tot_ft'], NF)

def _tnw(wb, data, pks, labels):
    ws = wb.create_sheet("TNW"); nc = 1+len(pks); _widths(ws, [45]+[18]*len(pks))
    _hdr_row(ws, 1, ["COMPONENT"]+[l.upper() for l in labels])
    tnw = data.get('tnw_analysis', {}); calc = tnw.get('calculation', {}); r = 2
    _sec_row(ws, r, "TNW CALCULATION", nc); r += 1
    _row(ws, r, ["Original TNW"]+[calc.get(pk,{}).get('original_tnw',0) for pk in pks], True, nfmt=NF); r += 1
    _row(ws, r, ["Less: Adjustments"]+[None]*len(pks), True, fc=C['err_ft']); r += 1
    for ak, al in [('less_intangibles','Intangible Assets'),('less_due_from_directors','Due from Directors'),
                   ('less_due_from_related_companies','Due from Related Co')]:
        _row(ws, r, [al]+[calc.get(pk,{}).get('adjustments',{}).get(ak,0) for pk in pks], indent=1, nfmt=NF); r += 1
    _row(ws, r, ["Total Adjustments"]+[calc.get(pk,{}).get('adjustments',{}).get('total_adjustments',0) for pk in pks], True, C['err_bg'], C['err_ft'], NF); r += 1
    _row(ws, r, ["Adjusted TNW"]+[calc.get(pk,{}).get('adjusted_tnw',0) for pk in pks], True, C['tot_bg'], C['tot_ft'], NF); r += 2
    notes = tnw.get('assessment', {}).get('notes', '')
    if notes:
        _sec_row(ws, r, "TNW ASSESSMENT", nc); r += 1
        c = ws.cell(r, 1, notes); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        ws.row_dimensions[r].height = 60

def _obs(wb, data, pks, labels):
    ws = wb.create_sheet("Observations"); _widths(ws, [25, 65, 12])
    summary = data.get('analysis_summary', {}); r = 1
    _hdr_row(ws, r, ["TOPIC", "OBSERVATION", ""]); r += 1
    for k, txt in summary.get('key_observations', {}).items():
        ws.cell(r, 1, k.replace('_',' ').title()).font = _ft(True)
        c = ws.cell(r, 2, txt); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.cell(r,1).border = BDR; c.border = BDR; ws.row_dimensions[r].height = 60; r += 1
    r += 1; _hdr_row(ws, r, ["POSITIVE INDICATORS", "DESCRIPTION", ""]); r += 1
    for it in summary.get('positive_indicators', []):
        ws.cell(r, 1, it.get('title','')).font = _ft(True, color=C['tot_ft'])
        c = ws.cell(r, 2, it.get('description','')); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.cell(r,1).border = BDR; c.border = BDR; ws.row_dimensions[r].height = 45; r += 1
    r += 1; _hdr_row(ws, r, ["AREAS OF CONCERN", "DESCRIPTION", "SEVERITY"]); r += 1
    for it in summary.get('areas_of_concern', []):
        sev = it.get('severity','')
        sc = C['err_ft'] if sev in ('CRITICAL','HIGH') else C['wrn_ft'] if sev=='MEDIUM' else '000000'
        ws.cell(r, 1, it.get('title','')).font = _ft(True, color=sc)
        c = ws.cell(r, 2, it.get('description','')); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.cell(r, 3, sev).font = _ft(True, color=sc)
        for i in range(1,4): ws.cell(r,i).border = BDR
        ws.row_dimensions[r].height = 45; r += 1
    r += 1; _hdr_row(ws, r, ["RECOMMENDATIONS", "ACTION", "PRIORITY"]); r += 1
    for it in summary.get('recommendations', []):
        ws.cell(r, 1, it.get('area','')).font = _ft(True)
        c = ws.cell(r, 2, it.get('action','')); c.font = _ft(sz=9); c.alignment = _al(wrap=True)
        ws.cell(r, 3, it.get('priority','')).font = _ft(True)
        for i in range(1,4): ws.cell(r,i).border = BDR
        ws.row_dimensions[r].height = 45; r += 1

def convert_json_to_excel(data: Dict) -> bytes:
    pks = _pks(data); labels = _labels(data)
    wb = Workbook()
    _summary(wb, data, pks, labels)
    _pl(wb, data, pks, labels)
    _bs(wb, data, pks, labels)
    _ratios(wb, data, pks, labels)
    _wc(wb, data, pks, labels)
    _dscr(wb, data, pks, labels)
    _tnw(wb, data, pks, labels)
    _obs(wb, data, pks, labels)
    for sn in ["P&L","Balance Sheet","Ratios","Working Capital","DSCR & Funding","TNW"]:
        if sn in wb.sheetnames: wb[sn].freeze_panes = 'B2'
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
